"""
Excel 報告生成器
從數據庫讀取測試記錄，生成 Excel (.xlsx) 格式的測試報告
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import config

from core.db_manager import get_session_actions, get_session
from i18n import t, get_lang, set_lang


def _action_type(key: str) -> str:
    """翻譯 action_type"""
    return t(f"action_{key}")


# ============================================================
# 樣式定義
# ============================================================

HEADER_FONT = Font(name='Microsoft JhengHei', bold=True, size=14, color='FFFFFF')
SUB_HEADER_FONT = Font(name='Microsoft JhengHei', bold=True, size=11, color='1A237E')
TABLE_HEADER_FONT = Font(name='Microsoft JhengHei', bold=True, size=11, color='FFFFFF')
NORMAL_FONT = Font(name='Microsoft JhengHei', size=10)
CODE_FONT = Font(name='Consolas', size=10)
TITLE_FILL = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
TABLE_HEADER_FILL = PatternFill(start_color='3949AB', end_color='3949AB', fill_type='solid')
ALT_ROW_FILL = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0'),
)


def _apply_table_style(ws, start_row, end_row, col_count):
    """對表格區域套用邊框和對齊"""
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=col_count):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)


def generate_excel_report(session_id: int, output_path: str = None, lang: str = None) -> str:
    """
    生成 Excel (.xlsx) 測試報告

    Args:
        session_id: 測試 session ID
        output_path: 輸出路徑，如果為 None 則默認為 reports/test_plan_{session_id}.xlsx
        lang: 語言覆蓋 ('zh' or 'en')，預設使用 config.LANGUAGE

    Returns:
        生成的 Excel 文件路徑
    """
    # 設定語言
    orig_lang = get_lang()
    if lang and lang in ("zh", "en"):
        set_lang(lang)
    elif config.LANGUAGE and config.LANGUAGE in ("zh", "en"):
        set_lang(config.LANGUAGE)

    actions = get_session_actions(session_id)
    session_row = get_session(session_id)

    if not session_row:
        raise ValueError(f"Session not found: {session_id}")

    session_info = {
        "id": session_row[0],
        "name": session_row[1],
        "url": session_row[2],
        "started_at": session_row[3],
        "ended_at": session_row[4],
        "status": session_row[5]
    }

    wb = Workbook()
    ws = wb.active
    ws.title = t("test_steps_detail")

    # ============================================================
    # 標題區
    # ============================================================
    ws.merge_cells('A1:K1')
    ws['A1'] = f"{t('report_title')} - {session_info['name']}"
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # 生成時間
    gen_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.merge_cells('A2:K2')
    ws['A2'] = f"{t('generated_at')}: {gen_time}"
    ws['A2'].font = Font(name='Microsoft JhengHei', size=9, color='78909C')
    ws['A2'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[2].height = 20

    # ============================================================
    # 測試概覽
    # ============================================================
    overview_start = 4
    ws.merge_cells(f'A{overview_start}:K{overview_start}')
    ws[f'A{overview_start}'] = f"📊 {t('test_overview')}"
    ws[f'A{overview_start}'].font = SUB_HEADER_FONT
    ws.row_dimensions[overview_start].height = 24

    overview_data = [
        (t('session_id'), str(session_info['id'])),
        (t('test_name'), session_info['name']),
        (t('test_url'), session_info['url']),
        (t('start_time'), session_info['started_at']),
        (t('end_time'), session_info['ended_at'] or '-'),
        (t('status'), session_info['status']),
        (t('total_actions'), str(len(actions))),
        (t('viewport_size'), f"{config.VIEWPORT_WIDTH} × {config.VIEWPORT_HEIGHT}"),
    ]

    for idx, (label, value) in enumerate(overview_data):
        row = overview_start + 1 + idx
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(name='Microsoft JhengHei', size=10, color='78909C')
        ws[f'B{row}'] = value
        ws[f'B{row}'].font = Font(name='Microsoft JhengHei', size=10, bold=True)
        ws.merge_cells(f'B{row}:K{row}')

    # ============================================================
    # 測試步驟詳情
    # ============================================================
    steps_header_row = overview_start + 1 + len(overview_data) + 1
    ws.merge_cells(f'A{steps_header_row}:K{steps_header_row}')
    ws[f'A{steps_header_row}'] = f"📝 {t('test_steps_detail')}"
    ws[f'A{steps_header_row}'].font = SUB_HEADER_FONT
    ws.row_dimensions[steps_header_row].height = 24

    # 表頭
    table_start = steps_header_row + 1
    headers = [
        t('step_num'),        # A
        t('time'),             # B
        t('action_type'),      # C
        t('page_title_col'),   # D
        t('element_id'),       # E
        t('element_name'),     # F
        t('element_type'),     # G
        t('target_selector'),  # H
        t('input_value'),      # I
        t('screenshot'),       # J
        t('purpose'),          # K
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start, column=col_idx, value=header)
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[table_start].height = 28

    # 設定欄寬
    col_widths = [8, 24, 14, 22, 22, 22, 12, 30, 20, 8, 36]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # 數據
    current_row = table_start + 1
    for i, action in enumerate(actions, 1):
        if len(action) >= 12:
            action_id, timestamp, action_type, selector, value, screenshot_path, purpose, page_title, page_url, element_name, element_type, element_id = action
        elif len(action) >= 11:
            action_id, timestamp, action_type, selector, value, screenshot_path, purpose, page_title, page_url, element_name, element_type = action
            element_id = ""
        else:
            action_id, timestamp, action_type, selector, value, screenshot_path, purpose = action
            page_title, page_url, element_name, element_type, element_id = "", "", "", "", ""

        zh_action = _action_type(action_type)
        screenshot_mark = "📷" if screenshot_path and os.path.exists(screenshot_path) else "-"

        row_data = [
            i,
            timestamp,
            zh_action,
            page_title or '-',
            element_id or '-',
            element_name or '-',
            element_type or '-',
            selector or '-',
            value or '-',
            screenshot_mark,
            purpose or '-',
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = CODE_FONT if col_idx in (8, 9, 5) else NORMAL_FONT
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        # 奇偶行交替顏色
        if i % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col_idx).fill = ALT_ROW_FILL

        ws.row_dimensions[current_row].height = 22
        current_row += 1

    # 套用邊框到整張表
    _apply_table_style(ws, table_start, current_row - 1, len(headers))

    # ============================================================
    # 自動驗收標準 (Sheet 2)
    # ============================================================
    ws2 = wb.create_sheet(title=t("auto_acceptance_criteria"))

    ws2.merge_cells('A1:D1')
    ws2['A1'] = f"✅ {t('auto_acceptance_criteria')}"
    ws2['A1'].font = HEADER_FONT
    ws2['A1'].fill = TITLE_FILL
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 36

    criteria_headers = [t('item_num'), t('verify_item'), t('verify_method'), t('expected_result')]
    criteria_widths = [8, 28, 28, 50]

    for col_idx, header in enumerate(criteria_headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, width in enumerate(criteria_widths, 1):
        ws2.column_dimensions[get_column_letter(idx)].width = width

    ws2.row_dimensions[3].height = 28

    criteria_row = 4
    for i, action in enumerate(actions, 1):
        if len(action) >= 12:
            action_id, timestamp, action_type, selector, value, screenshot_path, purpose, page_title, page_url, element_name, element_type, element_id = action
        elif len(action) >= 11:
            action_id, timestamp, action_type, selector, value, screenshot_path, purpose, page_title, page_url, element_name, element_type = action
        else:
            action_id, timestamp, action_type, selector, value, screenshot_path, purpose = action

        if action_type == "click":
            ws2.cell(row=criteria_row, column=1, value=i).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=2, value=t('click_verify')).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=3, value=t('click_desc')).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=4, value=t('click_result', selector=selector or '-')).font = NORMAL_FONT
            criteria_row += 1
        elif action_type == "type":
            ws2.cell(row=criteria_row, column=1, value=i).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=2, value=t('type_verify')).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=3, value=t('type_desc')).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=4, value=t('type_result', selector=selector or '-', value=value or '')).font = NORMAL_FONT
            criteria_row += 1
        elif action_type == "navigate":
            ws2.cell(row=criteria_row, column=1, value=i).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=2, value=t('navigate_verify')).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=3, value=t('navigate_desc')).font = NORMAL_FONT
            ws2.cell(row=criteria_row, column=4, value=t('navigate_result', selector=selector or '-')).font = NORMAL_FONT
            criteria_row += 1

    if criteria_row > 4:
        _apply_table_style(ws2, 3, criteria_row - 1, 4)

    # 頁尾
    footer_row = current_row + 1
    ws.merge_cells(f'A{footer_row}:K{footer_row}')
    ws[f'A{footer_row}'] = t('footer')
    ws[f'A{footer_row}'].font = Font(name='Microsoft JhengHei', size=9, color='90A4AE')
    ws[f'A{footer_row}'].alignment = Alignment(horizontal='center')

    # 儲存檔案
    if output_path is None:
        os.makedirs(config.REPORTS_DIR, exist_ok=True)
        output_path = os.path.join(config.REPORTS_DIR, f"test_plan_{session_id}.xlsx")

    wb.save(output_path)
    print(f"Excel Report: {output_path}")

    # 恢復原本語言
    set_lang(orig_lang)
    return output_path


if __name__ == "__main__":
    # 測試
    from core.db_manager import get_all_sessions
    sessions = get_all_sessions()
    if sessions:
        generate_excel_report(sessions[0][0])